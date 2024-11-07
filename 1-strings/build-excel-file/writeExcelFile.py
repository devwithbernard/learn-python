from wsgiref import headers

from openpyxl import Workbook

wb: Workbook = Workbook()

worksheet = wb.create_sheet('Comments')
worksheet.title = 'Comments'


def write_columns_title(heads: list[str]) -> None:
    """
    Writes columns titles.
    :param heads: list of column names
    :return: None
    """
    for i in range(len(heads)):
        head = heads[i]
        worksheet.cell(row=1, column=i + 1).value = head


def write_file(comments: list[dict], filename: str) -> None:
    """
    Writes comments to excel file
    :param comments: list of comments
    :param filename: name of excel file
    :return: None
    """
    h: list[str] = ['id', 'name', 'email', 'body']
    write_columns_title(h)
    for i in range(len(comments)):
        comment: dict = comments[i]
        worksheet.cell(row=i + 2, column=1).value = comment['id']
        worksheet.cell(row=i + 2, column=2).value = comment['name']
        worksheet.cell(row=i + 2, column=3).value = comment['email']
        worksheet.cell(row=i + 2, column=4).value = comment['body']

    # Saving file
    wb.save(filename)